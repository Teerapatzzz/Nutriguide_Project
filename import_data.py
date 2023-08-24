from openpyxl import load_workbook
import mysql.connector

# Excel
workbook = load_workbook('foodtable2.xlsx')
sheet = workbook.active
values = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    values.append(row)

# DATABASE
db = mysql.connector.connect(
    host='localhost',
    port=3306,
    user='root',
    password='Tt_110860',
    database='nutriv1'
)

cursor = db.cursor()
sql = """
    INSERT INTO foodtable (Food_ID, name_of_food_th, name_of_food_eng, kcal, water, protein, fat, carb,
    dietary_fiber, ash, calcium, phosphorus, magnesium, sodium, protassium, iron, copper, zinc, iodine,
    betacarotene, retinol, RAE, thiamin, riboflavin, niacin, vitamin_C, vitamin_E, sugar, vegan)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
"""

# Print SQL statement and data for debugging
print("SQL Statement:")
print(sql)
print("\nData:")
for row in values:
    print(row)

cursor.executemany(sql, values)
db.commit()
print('เพิ่มข้อมูล ' + str(cursor.rowcount) + ' แถว')
